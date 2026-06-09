"""
fornecedores_parser.py -- MKB-Dashboard
Importador do CADASTRO MESTRE DE FORNECEDORES do Protheus
(ex.: relatório "SA2 - Cadastro de Fornecedores" ou exportação equivalente).

Por que existe: o CSV de detalhe de despesas (CT2-Despesas) vem com o
campo "Razão Social" praticamente vazio (1 em 35.169 linhas) -- o nome
do fornecedor só pode ser extraído (de forma aproximada e truncada) do
texto do histórico do lançamento. Para exibir a RAZÃO SOCIAL OFICIAL na
análise de Despesas por Fornecedor, é preciso um cadastro separado que
relacione código (Cli_For/Lj) → razão social.

Este módulo grava esse relacionamento na tabela auxiliar
`fornecedores_cadastro` (ver schema.sql). Quando ela está vazia, a
análise funciona normalmente em modo "nome aproximado"; assim que for
populada, a tela passa a exibir o nome oficial -- sem precisar
reimportar o CT2 de despesas.

⚠ FORMATO AINDA NÃO CONFIRMADO -- este parser foi escrito de forma
defensiva (detecta colunas pelo nome, aceita CSV ou XLSX) porque o
arquivo real do cadastro ainda não foi enviado pelo usuário. Ajustar
`_RE_COL_CODIGO` / `_RE_COL_RAZAO` / `_RE_COL_FANTASIA` / `_RE_COL_DOC`
(e o delimitador/encoding do CSV, se necessário) ao receber o arquivo.
"""

import csv
import re
import sqlite3
from pathlib import Path

# ─── DETECÇÃO DE COLUNAS (por nome, case-insensitive) ───────────────────────
# O relatório do Protheus pode nomear as colunas de formas diferentes
# conforme o layout escolhido na exportação -- por isso usamos regex
# em vez de nomes fixos.

_RE_COL_CODIGO   = re.compile(r"c[oó]d.*(forn|cli)|cli_for|c[oó]digo$", re.IGNORECASE)
_RE_COL_RAZAO    = re.compile(r"raz[aã]o\s*social|nome\s*(do\s*)?(forn|cli)|raz[aã]o$", re.IGNORECASE)
_RE_COL_FANTASIA = re.compile(r"fantasia|apelido", re.IGNORECASE)
_RE_COL_DOC      = re.compile(r"cnpj|cpf|cgc", re.IGNORECASE)


def _limpa(s) -> str:
    if s is None:
        return ""
    return str(s).strip().strip("\xa0").strip()


def _mapear_colunas(header: list[str]) -> dict:
    """Localiza os índices das colunas relevantes a partir dos nomes do cabeçalho."""
    idx = {"codigo": None, "razao": None, "fantasia": None, "doc": None}
    for i, nome in enumerate(header):
        n = _limpa(nome)
        if idx["codigo"] is None and _RE_COL_CODIGO.search(n):
            idx["codigo"] = i
        elif idx["razao"] is None and _RE_COL_RAZAO.search(n):
            idx["razao"] = i
        elif idx["fantasia"] is None and _RE_COL_FANTASIA.search(n):
            idx["fantasia"] = i
        elif idx["doc"] is None and _RE_COL_DOC.search(n):
            idx["doc"] = i
    return idx


# ─── PARSER PRINCIPAL ────────────────────────────────────────────────────────

def parse_fornecedores(caminho: Path) -> list[dict]:
    """
    Lê o cadastro de fornecedores (CSV ou XLSX) e retorna lista de dicts:
        [{"cliente_cod": str, "razao_social": str,
          "nome_fantasia": str|None, "cnpj_cpf": str|None}, ...]

    Procura a linha de cabeçalho varrendo as primeiras ~10 linhas em busca
    de colunas reconhecíveis (código + razão social) -- tolera linhas de
    título de relatório antes do cabeçalho real, como nos CT2.
    """
    sufixo = caminho.suffix.lower()
    print(f"  Abrindo cadastro de fornecedores: {caminho.name}")

    if sufixo in (".xlsx", ".xlsm", ".xls"):
        linhas_brutas = _ler_xlsx(caminho)
    else:
        linhas_brutas = _ler_csv(caminho)

    if not linhas_brutas:
        print("  ⚠ Arquivo vazio ou ilegível.")
        return []

    # Procura o cabeçalho nas primeiras linhas
    header_idx, idx_cols = None, None
    for i, linha in enumerate(linhas_brutas[:10]):
        candidato = _mapear_colunas(linha)
        if candidato["codigo"] is not None and candidato["razao"] is not None:
            header_idx, idx_cols = i, candidato
            break

    if header_idx is None:
        print("  ✖ Não foi possível localizar colunas de código e razão social no cabeçalho.")
        print("    Ajuste _RE_COL_CODIGO / _RE_COL_RAZAO em fornecedores_parser.py "
              "para o layout real do arquivo.")
        return []

    registros, vistos, n_sem_codigo, n_sem_razao = [], set(), 0, 0
    for linha in linhas_brutas[header_idx + 1:]:
        if len(linha) <= max(idx_cols["codigo"], idx_cols["razao"]):
            continue

        cod   = _limpa(linha[idx_cols["codigo"]])
        razao = _limpa(linha[idx_cols["razao"]])
        if not cod:
            n_sem_codigo += 1
            continue
        if not razao:
            n_sem_razao += 1
            continue
        if cod in vistos:
            continue
        vistos.add(cod)

        fantasia = _limpa(linha[idx_cols["fantasia"]]) if idx_cols["fantasia"] is not None and len(linha) > idx_cols["fantasia"] else ""
        doc      = _limpa(linha[idx_cols["doc"]])      if idx_cols["doc"]      is not None and len(linha) > idx_cols["doc"]      else ""

        registros.append({
            "cliente_cod":   cod,
            "razao_social":  razao,
            "nome_fantasia": fantasia or None,
            "cnpj_cpf":      doc or None,
        })

    print(f"  {len(registros)} fornecedores | "
          f"{n_sem_codigo} sem código | {n_sem_razao} sem razão social | "
          f"cabeçalho na linha {header_idx + 1}")
    return registros


def _ler_csv(caminho: Path) -> list[list[str]]:
    """Lê CSV tolerando ';' ou ',' como delimitador e encoding latin-1/utf-8."""
    for encoding in ("latin-1", "utf-8-sig", "utf-8"):
        try:
            with open(caminho, encoding=encoding) as f:
                texto = f.read()
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        return []

    delimitador = ";" if texto.count(";") > texto.count(",") else ","
    reader = csv.reader(texto.splitlines(), delimiter=delimitador)
    return [[_limpa(c) for c in linha] for linha in reader if linha]


def _ler_xlsx(caminho: Path) -> list[list[str]]:
    import openpyxl
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas = [
        [_limpa(c) for c in row]
        for row in ws.iter_rows(values_only=True)
        if any(c is not None for c in row)
    ]
    wb.close()
    return linhas


# ─── PERSISTÊNCIA / IMPORTAÇÃO ───────────────────────────────────────────────

def salvar_fornecedores(conn: sqlite3.Connection, registros: list[dict]) -> int:
    """Upsert no cadastro mestre -- chave: cliente_cod."""
    if not registros:
        return 0
    conn.executemany(
        """
        INSERT INTO fornecedores_cadastro (cliente_cod, razao_social, nome_fantasia, cnpj_cpf)
        VALUES (:cliente_cod, :razao_social, :nome_fantasia, :cnpj_cpf)
        ON CONFLICT (cliente_cod) DO UPDATE SET
            razao_social  = excluded.razao_social,
            nome_fantasia = excluded.nome_fantasia,
            cnpj_cpf      = excluded.cnpj_cpf
        """,
        registros,
    )
    conn.commit()
    return len(registros)


def importar_fornecedores(caminho: Path, conn: sqlite3.Connection) -> dict:
    """Importa o cadastro de fornecedores para `fornecedores_cadastro`."""
    registros = parse_fornecedores(caminho)
    if not registros:
        return {"erro": "Nenhum fornecedor válido encontrado no arquivo "
                        "(verifique o layout -- ver _RE_COL_* em fornecedores_parser.py)."}
    qtd = salvar_fornecedores(conn, registros)
    return {"registros": qtd}


# ─── CLI DE TESTE ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from ingestion import get_conn, criar_schema

    if len(sys.argv) < 2:
        print("Uso: python fornecedores_parser.py <caminho_do_arquivo>")
        sys.exit(1)

    caminho = Path(sys.argv[1])
    conn = get_conn()
    criar_schema(conn)   # garante que fornecedores_cadastro existe
    resultado = importar_fornecedores(caminho, conn)
    conn.close()
    print("\nResultado:", resultado)
