"""
balancete_parser.py — MKB-Dashboard

Lê o Balancete de Verificação do Protheus (aba "12-00 - Balancete de Verificac")
e persiste o saldo atual (acumulado) por conta, para validar a DRE detalhada
contra o balancete e localizar diferenças.

Layout (linha 1 = cabeçalho, dados a partir da linha 2):
  col1 Conta | col2 Descricao | col3 Saldo anterior | col4 Debito |
  col5 Credito | col6 Mov periodo | col7 Saldo atual (magnitude) | col8 D/C

Convenção de sinal (igual à DRE / v_lancamentos): crédito = +, débito = -.
  Saldo atual assinado = +magnitude se 'C', -magnitude se 'D'.
"""

import sqlite3
from pathlib import Path

import openpyxl

from config import EMPRESAS


def _num(v) -> float:
    """Converte célula numérica ou texto pt-BR para float (magnitude)."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    # remove sufixo de sinal se vier no próprio texto
    if s[-1] in ("C", "D"):
        s = s[:-1].strip()
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_signed_text(cell) -> float:
    """'136.579,05 C' -> +136579.05 ; '... D' -> negativo ; vazio -> 0.0"""
    if cell is None:
        return 0.0
    if isinstance(cell, (int, float)):
        return float(cell)
    s = str(cell).strip()
    if not s:
        return 0.0
    sign = 1
    if s.endswith("D"):
        sign = -1
    mag = _num(s)
    return sign * abs(mag)


def _saldo_assinado(mag_cell, sign_cell) -> float:
    """Saldo atual assinado (convenção DRE: C = +, D = -).
    O sinal pode vir de 3 formas, dependendo do layout do balancete:
      (a) coluna D/C separada (sign_cell);
      (b) sufixo no próprio texto do saldo ('135.000,00 D');
      (c) sinal do próprio número (saldo já negativo).
    """
    sign = (str(sign_cell).strip().upper() if sign_cell is not None else "")
    # (b) sufixo embutido no texto do saldo
    if sign not in ("D", "C") and isinstance(mag_cell, str):
        t = mag_cell.strip().upper()
        if t.endswith("D"):
            sign = "D"
        elif t.endswith("C"):
            sign = "C"
    val = _num(mag_cell)
    if sign == "D":
        return -abs(val)
    if sign == "C":
        return abs(val)
    # (c) sem indicador D/C → usa o sinal do próprio número
    return val


def parse_balancete(caminho: Path, empresa_id: int) -> list[dict]:
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)

    # Aba do balancete (começa com "12-00 - Balancete...") ou a 1ª aba
    aba = next((s for s in wb.sheetnames if "balancete" in s.lower()), wb.sheetnames[0])
    ws = wb[aba]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    # Localiza as colunas pelo cabeçalho (robusto a layouts diferentes entre
    # empresas/meses). A coluna de sinal D/C costuma ser a imediatamente após
    # "Saldo atual" (cabeçalho em branco).
    header = [str(c or "").strip().lower() for c in rows[0]]
    def _col(*chaves, default=None):
        for idx, h in enumerate(header):
            if any(k in h for k in chaves):
                return idx
        return default
    i_conta = _col("conta", default=0)
    i_desc  = _col("descric", default=1)
    i_mov   = _col("mov", default=5)
    i_saldo = _col("saldo atual", "saldo final", default=6)
    i_sign  = i_saldo + 1   # coluna D/C logo após o saldo

    registros: list[dict] = []
    for i, row in enumerate(rows):
        if i == 0:
            continue  # cabeçalho
        if not row or len(row) <= i_conta or row[i_conta] is None:
            continue
        conta = str(row[i_conta]).strip()
        if not conta or not conta[0].isdigit():
            continue
        descricao   = str((row[i_desc] if len(row) > i_desc else "") or "").strip()
        mov_periodo = _parse_signed_text(row[i_mov]) if len(row) > i_mov else 0.0
        saldo_atual = _saldo_assinado(
            row[i_saldo] if len(row) > i_saldo else None,
            row[i_sign] if len(row) > i_sign else None,
        )
        registros.append({
            "empresa_id":  empresa_id,
            "conta_cod":   conta,
            "descricao":   descricao,
            "saldo_atual": round(saldo_atual, 2),
            "mov_periodo": round(mov_periodo, 2),
        })

    wb.close()
    return registros


def salvar_balancete(conn: sqlite3.Connection, registros: list[dict],
                     competencia: str) -> int:
    """Substitui o balancete da (empresa, competência) pelos novos registros."""
    if not registros:
        return 0
    emp_id = registros[0]["empresa_id"]
    conn.execute(
        "DELETE FROM balancete WHERE empresa_id=? AND competencia=?",
        (emp_id, competencia)
    )
    for r in registros:
        r["competencia"] = competencia
    conn.executemany(
        """
        INSERT INTO balancete
            (empresa_id, competencia, conta_cod, descricao, saldo_atual, mov_periodo)
        VALUES
            (:empresa_id, :competencia, :conta_cod, :descricao, :saldo_atual, :mov_periodo)
        """,
        registros,
    )
    conn.commit()
    return len(registros)


def importar_balancete(caminho: Path, empresa_chave: str, competencia: str,
                       conn: sqlite3.Connection) -> dict:
    emp = EMPRESAS.get(empresa_chave)
    if not emp:
        return {"erro": f"Empresa '{empresa_chave}' não encontrada"}
    registros = parse_balancete(caminho, emp["id"])
    qtd = salvar_balancete(conn, registros, competencia)
    return {"registros": qtd, "empresa": emp["sigla"], "competencia": competencia}
