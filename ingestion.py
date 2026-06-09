"""
ingestion.py — MKB-Dashboard  Sprint 1
Lê os arquivos Excel do Protheus (Template DRE Protheus) e persiste
os lançamentos no banco SQLite.

Uso direto:
    python ingestion.py --ano 2026 --mes 4
    python ingestion.py --ano 2026 --mes 1 4   # importa jan até abr

Formato dos valores no Protheus:
    "      1.234.567,89 C"  →  +1234567.89  (crédito = receita)
    "      1.234.567,89 D"  →  -1234567.89  (débito   = despesa)
    "                   -"  →   0.00         (zero)
    ""                       →   0.00         (vazio)
"""

import re
import sqlite3
import argparse
from pathlib import Path
from datetime import datetime

import openpyxl

from config import (
    DB_PATH, EMPRESAS, SHEET_TEMPLATE,
    HEADER_ROW, COL_COD_CONTA, COL_DESCRICAO, COL_PERIODO_INI, COL_PERIODO_FIM,
    SKIP_DESCRICOES, caminho_mkb, caminho_gnileb
)


# ─── BANCO ────────────────────────────────────────────────────────────────────

def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def criar_schema(conn: sqlite3.Connection) -> None:
    schema = (Path(__file__).parent / "schema.sql").read_text(encoding="utf-8")
    # Dropa a view antes de recriar (ALTER VIEW não existe no SQLite)
    try:
        conn.execute("DROP VIEW IF EXISTS v_lancamentos")
    except Exception:
        pass
    conn.executescript(schema)

    # Migração: bancos criados antes da coluna `parceiro_cod` (CREATE TABLE
    # IF NOT EXISTS não adiciona colunas em tabela já existente)
    cols = {row[1] for row in conn.execute("PRAGMA table_info(razao)").fetchall()}
    if "parceiro_cod" not in cols:
        conn.execute("ALTER TABLE razao ADD COLUMN parceiro_cod TEXT")

    conn.commit()


def seed_empresas(conn: sqlite3.Connection) -> None:
    for chave, emp in EMPRESAS.items():
        conn.execute(
            "INSERT OR IGNORE INTO empresas (id, sigla, nome, cnpj) VALUES (?,?,?,?)",
            (emp["id"], emp["sigla"], emp["nome"], emp["cnpj"])
        )
    conn.commit()


# ─── PARSER DO VALOR PROTHEUS ─────────────────────────────────────────────────

_RE_VALOR = re.compile(r"([\d.,]+)\s*([CD])\s*$")

def parse_valor(texto: str) -> float | None:
    """
    Converte o texto do Protheus para float.
    Retorna None se a célula estiver vazia ou for linha de total.

    Exemplos:
        "      1.234,56 C"  →  +1234.56
        "        314,55 D"  →  -314.55
        "      1.234,56  "  →  None  (sem C/D = linha de total ou fórmula)
        "                -"  →   0.0
        ""                  →  None (ignorar)
    """
    if not texto:
        return None
    t = texto.strip()
    if not t or t in ("-", "-   ", "0"):
        return 0.0

    m = _RE_VALOR.search(t)
    if not m:
        return None  # não tem C/D → linha de total ou separador

    num_str, sinal = m.group(1), m.group(2)
    # Remove pontos de milhar, troca vírgula decimal por ponto
    num_str = num_str.replace(".", "").replace(",", ".")
    try:
        valor = float(num_str)
    except ValueError:
        return None

    return valor if sinal == "C" else -valor


# ─── PARSER DO CABEÇALHO DE PERÍODO ───────────────────────────────────────────

_RE_PERIODO_DATA = re.compile(r"PERIODO\s+\d+\s+(\d{2})/(\d{2})\s*-")

def parse_competencia(cabecalho: str, ano_fallback: int, numero_periodo: int) -> str | None:
    """
    Extrai 'YYYY-MM' do cabeçalho de coluna do Protheus.

    Casos:
        "PERIODO 1 01/01 - 31/01"  →  "{ano_fallback}-01"
        "PERIODO 2 01/02 - 28/02"  →  "{ano_fallback}-02"
        "PERIODO 5"                →  None  (sem dados ainda)
    """
    if not cabecalho:
        return None
    m = _RE_PERIODO_DATA.search(cabecalho)
    if m:
        # dd/mm → usa apenas o mês
        mes = int(m.group(2))
        return f"{ano_fallback}-{mes:02d}"
    # Sem data no cabeçalho = período sem dados
    return None


# ─── LEITURA DO EXCEL ─────────────────────────────────────────────────────────

def ler_template_dre(
    caminho: Path,
    empresa_chave: str,
    ano: int,
) -> list[dict]:
    """
    Lê a aba 'Template DRE Protheus' do arquivo Excel indicado.
    Retorna lista de dicts: {empresa_id, competencia, conta_cod, descricao, valor}
    """
    emp       = EMPRESAS[empresa_chave]
    header_row = HEADER_ROW[empresa_chave]

    print(f"  Abrindo: {caminho.name}")
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)

    if SHEET_TEMPLATE not in wb.sheetnames:
        print(f"  ⚠  Aba '{SHEET_TEMPLATE}' não encontrada em {caminho.name}")
        wb.close()
        return []

    ws = wb[SHEET_TEMPLATE]
    rows = list(ws.iter_rows(values_only=True))

    # Cabeçalho: linha header_row (1-based → índice header_row - 1)
    cabecalho = rows[header_row - 1]

    # Mapear número de período → competência
    # Coluna 1=Cod, 2=Desc, 3..14=PERIODO 1..12 (índices 0..1, 2..13)
    periodo_comp: dict[int, str] = {}
    for col_idx in range(COL_PERIODO_INI - 1, COL_PERIODO_FIM):   # índices 2..13
        texto_cabecalho = str(cabecalho[col_idx] or "").strip()
        numero_periodo  = col_idx - (COL_PERIODO_INI - 1) + 1     # 1..12
        comp = parse_competencia(texto_cabecalho, ano, numero_periodo)
        if comp:
            periodo_comp[col_idx] = comp

    if not periodo_comp:
        print(f"  ⚠  Nenhum período com data encontrado em {caminho.name}")
        wb.close()
        return []

    competencias_encontradas = sorted(set(periodo_comp.values()))
    print(f"  Competências: {', '.join(competencias_encontradas)}")

    registros: list[dict] = []
    # Dados: a partir da linha seguinte ao cabeçalho
    for row in rows[header_row:]:
        cod_conta  = str(row[COL_COD_CONTA - 1] or "").strip()
        descricao  = str(row[COL_DESCRICAO - 1] or "").strip()

        # Ignora linhas de total, vazias ou sem código de conta
        if not cod_conta or descricao.upper() in SKIP_DESCRICOES:
            continue
        # Código de conta deve começar com dígito (3.x, 4.x, etc.)
        if not cod_conta[0].isdigit():
            continue

        for col_idx, competencia in periodo_comp.items():
            texto_celula = str(row[col_idx] or "").strip()
            valor = parse_valor(texto_celula)
            if valor is None:
                continue  # célula sem C/D (total ou vazia sem formato)

            registros.append({
                "empresa_id":  emp["id"],
                "competencia": competencia,
                "conta_cod":   cod_conta,
                "descricao":   descricao,
                "valor":       valor,
            })

    wb.close()
    print(f"  {len(registros)} registros lidos.")
    return registros


# ─── PERSISTÊNCIA ─────────────────────────────────────────────────────────────

def salvar_lancamentos(
    conn: sqlite3.Connection,
    registros: list[dict],
    arquivo: Path,
) -> int:
    """
    Upsert dos lançamentos.
    Registros do mesmo (empresa, competencia, conta) são substituídos.
    """
    if not registros:
        return 0

    conn.executemany(
        """
        INSERT INTO lancamentos (empresa_id, competencia, conta_cod, valor)
        VALUES (:empresa_id, :competencia, :conta_cod, :valor)
        ON CONFLICT (empresa_id, competencia, conta_cod)
        DO UPDATE SET valor = excluded.valor
        """,
        registros,
    )

    # Salvar contas novas (ordem: cod, empresa_id, descricao)
    contas_unicas = {
        (r["conta_cod"], r["empresa_id"], r["descricao"])
        for r in registros
    }
    conn.executemany(
        """
        INSERT OR IGNORE INTO contas (cod, empresa_id, descricao)
        VALUES (?, ?, ?)
        """,
        contas_unicas,
    )

    # Log de importação
    empresa_id   = registros[0]["empresa_id"]
    competencias = sorted({r["competencia"] for r in registros})
    for comp in competencias:
        qtd = sum(1 for r in registros if r["competencia"] == comp)
        conn.execute(
            """
            INSERT INTO importacoes (empresa_id, competencia, arquivo, registros)
            VALUES (?, ?, ?, ?)
            """,
            (empresa_id, comp, str(arquivo), qtd)
        )

    conn.commit()
    return len(registros)


# ─── FUNÇÃO PRINCIPAL DE IMPORTAÇÃO ──────────────────────────────────────────

def importar(ano: int, mes: int, empresas: list | None = None) -> dict:
    """
    Importa dados do Protheus para o mês/ano indicado.
    empresas: lista de chaves a importar, ex: ['mkb', 'gnileb'] (default: ambas).
    Retorna resumo da operação.
    """
    if empresas is None:
        empresas = ["mkb", "gnileb"]

    conn = get_conn()
    criar_schema(conn)
    seed_empresas(conn)

    resultado = {"ano": ano, "mes": mes, "mkb": 0, "gnileb": 0, "erros": []}

    # ── MKB ──────────────────────────────────────────────────────────────────
    if "mkb" in empresas:
        arq_mkb = caminho_mkb(ano, mes)
        if arq_mkb.exists():
            print(f"\n[MKB] {arq_mkb.name}")
            try:
                regs = ler_template_dre(arq_mkb, "mkb", ano)
                resultado["mkb"] = salvar_lancamentos(conn, regs, arq_mkb)
            except Exception as e:
                msg = f"MKB {ano}/{mes:02d}: {e}"
                resultado["erros"].append(msg)
                print(f"  ✖ {msg}")
        else:
            msg = f"Arquivo MKB não encontrado: {arq_mkb}"
            resultado["erros"].append(msg)
            print(f"  ⚠ {msg}")

    # ── GNILEB ───────────────────────────────────────────────────────────────
    if "gnileb" in empresas:
        arq_gnileb = caminho_gnileb(ano, mes)
        if arq_gnileb.exists():
            print(f"\n[GNILEB] {arq_gnileb.name}")
            try:
                regs = ler_template_dre(arq_gnileb, "gnileb", ano)
                resultado["gnileb"] = salvar_lancamentos(conn, regs, arq_gnileb)
            except Exception as e:
                msg = f"GNILEB {ano}/{mes:02d}: {e}"
                resultado["erros"].append(msg)
                print(f"  ✖ {msg}")
        else:
            msg = f"Arquivo GNILEB não encontrado: {arq_gnileb}"
            resultado["erros"].append(msg)
            print(f"  ⚠ {msg}")

    conn.close()
    return resultado


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Importa dados do Protheus para o banco MKB-Dashboard")
    parser.add_argument("--ano", type=int, default=datetime.now().year)
    parser.add_argument("--mes", type=int, nargs="+", default=[datetime.now().month],
                        help="Um ou mais meses (1-12). Ex: --mes 1 2 3 4")
    args = parser.parse_args()

    meses = args.mes if isinstance(args.mes, list) else [args.mes]
    for mes in meses:
        print(f"\n{'='*60}")
        print(f"Importando {args.ano}/{mes:02d}")
        print('='*60)
        res = importar(args.ano, mes)
        print(f"\nResultado: MKB={res['mkb']} registros | GNILEB={res['gnileb']} registros")
        if res["erros"]:
            for e in res["erros"]:
                print(f"  ✖ {e}")


if __name__ == "__main__":
    main()
