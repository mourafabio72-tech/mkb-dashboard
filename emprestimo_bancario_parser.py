"""
emprestimo_bancario_parser.py -- MKB-Dashboard
Importador do CRONOGRAMA DE AMORTIZAÇÃO (tabela Price) de um empréstimo
bancário, a partir da planilha de simulação/controle do próprio banco.

Usado como fonte do detalhamento mês a mês (saldo devedor, valor da parcela,
amortização, juros) enquanto o Razão da empresa não tem as 4 contas do
empréstimo (ver emprestimos_bancarios em app.py) -- uma vez que o Razão
cobrir essas contas, a rota /endividamento-bancario passa a priorizar o
saldo real do Razão sobre este cronograma.

Formato esperado: uma aba contendo, na linha de cabeçalho, as colunas
N, Data, Amort[ização], Juros, SD (saldo devedor), PMT (valor da parcela)
-- nessa ordem ou não, com ou sem acento. A aba é localizada automaticamente
entre as do arquivo (a planilha do banco costuma ter várias abas auxiliares).
A linha N=0 (mês do desembolso, ainda não é parcela) é ignorada.
"""

import re
import unicodedata
import sqlite3
from pathlib import Path

import openpyxl


def _normaliza(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()


def _achar_aba_cronograma(wb):
    """Procura a 1ª aba cuja linha de cabeçalho tem N + Data + Amort*."""
    for nome in wb.sheetnames:
        ws = wb[nome]
        for linha in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            normalizados = [_normaliza(c) for c in linha]
            if "n" in normalizados and "data" in normalizados and any(c.startswith("amort") for c in normalizados):
                return nome
    return None


def _indices_colunas(header_norm: list) -> dict:
    def idx(*opcoes):
        for i, c in enumerate(header_norm):
            if c and any(c == o or c.startswith(o) for o in opcoes):
                return i
        return None

    return {
        "n":     idx("n"),
        "data":  idx("data"),
        "amort": idx("amort"),
        "juros": idx("juros"),
        "sd":    idx("sd"),
        "pmt":   idx("pmt"),
    }


def parse_cronograma(caminho: Path) -> list[dict]:
    """
    Lê o cronograma de amortização e retorna lista de dicts:
        {numero_parcela, competencia, amortizacao, juros, saldo_devedor, valor_parcela}
    Ignora a linha N=0 (desembolso). Retorna [] se não achar a aba certa.
    """
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    nome_aba = _achar_aba_cronograma(wb)
    if not nome_aba:
        wb.close()
        return []

    ws = wb[nome_aba]
    registros = []
    idxs = None

    for row in ws.iter_rows(values_only=True):
        normalizados = [_normaliza(c) for c in row]

        if idxs is None:
            # ainda procurando a linha de cabeçalho dentro desta aba
            if "n" in normalizados and "data" in normalizados:
                idxs = _indices_colunas(normalizados)
            continue

        def col(chave):
            i = idxs.get(chave)
            return row[i] if i is not None and i < len(row) else None

        n_val = col("n")
        if not isinstance(n_val, (int, float)):
            continue
        n_val = int(n_val)
        if n_val == 0:
            continue  # mês do desembolso, ainda não é parcela

        data_val = col("data")
        if not data_val:
            continue
        m = re.match(r"(\d{1,2})/(\d{4})", str(data_val).strip())
        if not m:
            continue
        competencia = f"{m.group(2)}-{int(m.group(1)):02d}"

        registros.append({
            "numero_parcela": n_val,
            "competencia":    competencia,
            "amortizacao":    col("amort"),
            "juros":          col("juros"),
            "saldo_devedor":  col("sd"),
            "valor_parcela":  col("pmt"),
        })

    wb.close()
    registros.sort(key=lambda r: r["numero_parcela"])
    return registros


def importar_cronograma(caminho: Path, emprestimo_id: int, conn: sqlite3.Connection) -> dict:
    """Importa (substituindo) o cronograma de um empréstimo já cadastrado."""
    registros = parse_cronograma(caminho)
    if not registros:
        return {"erro": "Não encontrei uma aba com colunas N / Data / Amortização / Juros / SD / PMT nessa planilha."}

    conn.execute("DELETE FROM emprestimos_parcelas WHERE emprestimo_id=?", (emprestimo_id,))
    conn.executemany(
        """
        INSERT INTO emprestimos_parcelas
            (emprestimo_id, numero_parcela, competencia, amortizacao, juros, saldo_devedor, valor_parcela)
        VALUES (:emprestimo_id, :numero_parcela, :competencia, :amortizacao, :juros, :saldo_devedor, :valor_parcela)
        """,
        [{**r, "emprestimo_id": emprestimo_id} for r in registros],
    )
    conn.commit()
    return {"registros": len(registros), "competencia_ini": registros[0]["competencia"],
            "competencia_fim": registros[-1]["competencia"]}
