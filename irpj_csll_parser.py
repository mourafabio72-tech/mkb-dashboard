"""
irpj_csll_parser.py -- MKB-Dashboard
Importador da planilha ANUAL de apuração de IRPJ/CSLL (Lucro Real) -- arquivo
preenchido manualmente pelo contador (fora do Protheus), upload independente
do CT2.

Formato esperado (aba "ANUAL"):
  Coluna A = conta contábil (pode estar vazia em linhas de cálculo)
  Coluna B = descrição do item
  Colunas C em diante = um mês cada (Jan, Fev, ..., Dez), valores acumulados

A planilha traz 2 blocos na mesma aba, na ordem: CSLL primeiro, IRPJ depois
(estrutura LALUR/LACS: resultado contábil acumulado, despesas IRPJ/CSLL
contabilizadas, adições, exclusões, base de cálculo, compensação de prejuízo
fiscal, alíquota (CSLL 9% / IRPJ 15%+10%), retenções, valor final devido).

Esta importação é INDEPENDENTE do grupo "IRPJ_CSLL" calculado em dre_engine.py
a partir do CT2 (que é apenas a provisão contábil da conta 4.5.x) -- não cruza
com aquele dado.
"""

import re
import unicodedata
import sqlite3
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel

from config import EMPRESAS

SHEET_NAME = "ANUAL"

_MESES_PT = ("jan", "fev", "mar", "abr", "mai", "jun",
             "jul", "ago", "set", "out", "nov", "dez")

# Marcadores (tolerantes a acento/caixa) usados para classificar as linhas.
# "a pagar" foi removido: dá falso positivo em linhas que não são o total
# final (ex.: "Provisões de Serviços a Pagar", um item de adição/exclusão).
# "a recolher" é o marcador real observado na planilha do cliente
# ("CSLL A RECOLHER" / "IRPJ A RECOLHER" -- sempre a última linha do bloco).
_MARCADORES_DESTAQUE = ("a recolher", "final")

# Linhas de subtotal/marco do cálculo -- ficam em negrito na tela para
# evidenciar a estrutura da apuração (lucro contábil/líquido, adições,
# exclusões, compensação, base bruta, CSLL devida/IRPJ devido). São
# diferentes da linha "destaque" (valor final a recolher).
_SUBTOTAL_EXATO = (
    "adicoes", "adicoes permanente", "adicoes temporarias",
    "exclusoes", "exclusao permanente", "exclusoes temporarias",
    "csll devida", "irpj devido",
)
_SUBTOTAL_PREFIXO = ("lucro contabil", "lucro liquido")
_SUBTOTAL_CONTEM = ("compensacao", "base bruta")


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def _norm(s) -> str:
    """minúsculo, sem acento -- para comparação tolerante de texto."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def _is_mes_cell(texto: str) -> bool:
    t = _norm(texto)
    return any(t.startswith(m) for m in _MESES_PT)


def _is_subtotal_linha(desc_norm: str) -> bool:
    """Identifica linhas de marco/subtotal (negrito na tela). Usa igualdade
    exata para "Adições"/"Exclusões"/"...Devida(o)" -- evita pegar os itens
    de detalhe abaixo delas (ex.: "Confraternização", "Brindes")."""
    d = desc_norm.strip(" :-—.()")
    if d in _SUBTOTAL_EXATO:
        return True
    if any(d.startswith(p) for p in _SUBTOTAL_PREFIXO):
        return True
    if any(c in d for c in _SUBTOTAL_CONTEM):
        return True
    return False


_RE_DATA_NUMERICA = re.compile(r"^(\d{1,2})[/\-.](\d{2,4})$")   # "01/2026", "01-26"
_RE_DATA_ISO      = re.compile(r"^(\d{4})[/\-.](\d{1,2})$")      # "2026-01"

def _resolver_mes_ano(valor, ano_default: int) -> tuple[int, int] | None:
    """
    Resolve uma célula de cabeçalho (mês) para (mes, ano), aceitando:
      - objeto datetime/date (célula formatada como data no Excel)
      - número de série Excel "solto" (célula sem formatação de data aplicada
        -- o openpyxl só devolve datetime quando a célula TEM esse formato;
        sem ele, vem um float/int puro, ex.: 46023.0 em vez de 2026-01-01)
      - texto com nome do mês em português ("Jan/26", "Janeiro/2026", ...)
      - texto numérico ("01/2026", "01-26", "2026-01")
    Retorna None se não reconhecer o formato.
    """
    if isinstance(valor, (datetime, date)):
        return valor.month, valor.year

    # Número de série Excel sem formatação de data na célula (intervalo
    # plausível: ~2015 a ~2099 em dias desde a época do Excel)
    if isinstance(valor, (int, float)) and not isinstance(valor, bool) and 40000 <= valor <= 80000:
        try:
            dt = from_excel(valor)
            if dt is not None:
                return dt.month, dt.year
        except Exception:
            pass

    texto = str(valor).strip()
    if not texto:
        return None

    # Nome do mês em português + ano opcional
    if _is_mes_cell(texto):
        mes_num = next(
            (i + 1 for i, m in enumerate(_MESES_PT) if _norm(texto).startswith(m)),
            None,
        )
        if mes_num:
            m_ano = re.search(r"(\d{2,4})", texto)
            ano = ano_default
            if m_ano:
                a = int(m_ano.group(1))
                ano = 2000 + a if a < 100 else a
            return mes_num, ano

    # "mm/aaaa" ou "mm/aa"
    m = _RE_DATA_NUMERICA.match(texto)
    if m:
        mes_num, a = int(m.group(1)), int(m.group(2))
        if 1 <= mes_num <= 12:
            ano = 2000 + a if a < 100 else a
            return mes_num, ano

    # "aaaa-mm" (ISO)
    m = _RE_DATA_ISO.match(texto)
    if m:
        a, mes_num = int(m.group(1)), int(m.group(2))
        if 1 <= mes_num <= 12:
            return mes_num, a

    return None


def _to_float(valor) -> float | None:
    """Converte célula numérica ou texto BR ('1.234,56' / '(123,45)') para float.
    '-' (convenção contábil para "zero") vira 0.0 explícito -- diferente de
    célula genuinamente vazia, que vira None. Essa distinção importa para a
    linha "A Recolher": quando o imposto é zero (coberto por antecipações),
    o card deve mostrar "R$ 0", não "—" (que sugeriria falta de dado)."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip()
    if not s:
        return None
    if s in ("-", "--", "—"):
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(".", "").replace(",", ".")
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


# ─── LEITURA DO EXCEL ─────────────────────────────────────────────────────────

def parse_irpj_csll(caminho: Path) -> dict:
    """
    Lê a aba ANUAL (case-insensitive; fallback 1ª aba) e retorna:
        {"competencias": ["2026-01", ...], "registros": [...]}

    Cada registro: {competencia, secao, ordem, conta_cod, descricao, valor, is_destaque}
    `ordem` preserva a posição original da linha na planilha (1-based, contínua
    entre os dois blocos) para reconstrução fiel na tela.
    """
    print(f"  Abrindo IRPJ/CSLL (ANUAL): {caminho.name}")
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)

    nome_aba = next(
        (n for n in wb.sheetnames if n.strip().upper() == SHEET_NAME),
        wb.sheetnames[0],
    )
    ws = wb[nome_aba]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {"competencias": [], "registros": []}

    # 1. Detecta linha de cabeçalho com os meses (colunas C+, índice >= 2).
    #    Aceita texto ("Jan/26"), datas numéricas ("01/2026") e células
    #    formatadas como data (datetime/date) -- comum quando o Excel
    #    formata o cabeçalho do mês como data em vez de texto.
    ano_default = datetime.now().year
    header_idx = None
    col_comp: dict[int, str] = {}
    for r_idx, row in enumerate(rows[:15]):
        achou = {}
        nao_reconhecidas = []
        for c_idx in range(2, len(row)):
            cel = row[c_idx]
            if cel in (None, ""):
                continue
            resolvido = _resolver_mes_ano(cel, ano_default)
            if resolvido:
                mes_num, ano = resolvido
                achou[c_idx] = f"{ano}-{mes_num:02d}"
            else:
                nao_reconhecidas.append((c_idx, repr(cel), type(cel).__name__))
        if achou and nao_reconhecidas:
            # Diagnóstico: linha de cabeçalho identificada, mas com células
            # vizinhas não reconhecidas como mês -- ajuda a investigar casos
            # como número de série Excel fora do intervalo esperado.
            print(f"  AVISO: colunas de cabecalho nao reconhecidas como mes na linha {r_idx+1}: {nao_reconhecidas}")
        if achou:
            header_idx = r_idx
            col_comp = achou
            break

    if header_idx is None:
        print("  AVISO: nenhuma linha de cabecalho de mes encontrada na aba.")
        return {"competencias": [], "registros": []}

    cols_ordenadas = sorted(col_comp.keys())

    # 3. Processa linhas de dados (abaixo do cabeçalho), classificando seção
    secao_atual = "CSLL"
    registros = []
    ordem = 0
    for row in rows[header_idx + 1:]:
        conta_cod = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        descricao = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not descricao:
            continue

        ordem += 1
        desc_norm = _norm(descricao)
        desc_pontuacao_solta = desc_norm.strip(" :-—.")

        # Transição CSLL -> IRPJ: a planilha real usa uma linha-título isolada
        # contendo só a palavra "IRPJ" (sem mais nada) para abrir o 2º bloco
        # -- ex.: "IRPJ" sozinho na coluna B, seguido de "Descrição" repetido.
        # Mantém também o padrão alternativo "irpj"+"base" (outras planilhas
        # podem nomear o início do bloco como "Base de Cálculo do IRPJ").
        eh_titulo_irpj = desc_pontuacao_solta == "irpj"
        eh_base_irpj   = "irpj" in desc_norm and "base" in desc_norm
        if secao_atual == "CSLL" and (eh_titulo_irpj or eh_base_irpj):
            secao_atual = "IRPJ"

        # is_destaque/is_subtotal dependem só da DESCRIÇÃO da linha, nunca do
        # valor -- "CSLL/IRPJ A RECOLHER" continua sendo a linha-destaque
        # mesmo quando o valor é zero em todos os meses (ex.: quando as
        # antecipações cobrem 100% do imposto devido). Calcular em função do
        # valor faria o card de resumo "desaparecer" justamente no caso mais
        # comum (imposto zerado).
        is_destaque = any(m in desc_norm for m in _MARCADORES_DESTAQUE)
        is_subtotal = _is_subtotal_linha(desc_norm)

        valores_por_col = {
            c_idx: (_to_float(row[c_idx]) if c_idx < len(row) else None)
            for c_idx in cols_ordenadas
        }

        for c_idx, competencia in col_comp.items():
            registros.append({
                "competencia": competencia,
                "secao": secao_atual,
                "ordem": ordem,
                "conta_cod": conta_cod or None,
                "descricao": descricao,
                "valor": valores_por_col.get(c_idx),
                "is_destaque": 1 if is_destaque else 0,
                "is_subtotal": 1 if is_subtotal else 0,
            })

    competencias = sorted(set(col_comp.values()))
    print(f"  {len(registros)} linhas x competência | competências: {', '.join(competencias)}")

    return {"competencias": competencias, "registros": registros}


# ─── PERSISTÊNCIA ─────────────────────────────────────────────────────────────

def salvar_irpj_csll(conn: sqlite3.Connection, registros: list, empresa_id: int, arquivo: Path) -> int:
    """
    Substitui (delete + insert) os dados das competências presentes neste
    arquivo, para esta empresa -- evita lixo de linhas removidas/alteradas
    entre uploads do mesmo período (a planilha pode ganhar/perder linhas).
    """
    if not registros:
        return 0

    competencias = sorted({r["competencia"] for r in registros})
    for comp in competencias:
        conn.execute(
            "DELETE FROM irpj_csll WHERE empresa_id=? AND competencia=?",
            (empresa_id, comp),
        )

    conn.executemany(
        """
        INSERT INTO irpj_csll
            (empresa_id, competencia, secao, ordem, conta_cod, descricao, valor, is_destaque, is_subtotal)
        VALUES (:empresa_id, :competencia, :secao, :ordem, :conta_cod, :descricao, :valor, :is_destaque, :is_subtotal)
        """,
        [{**r, "empresa_id": empresa_id} for r in registros],
    )

    for comp in competencias:
        qtd = sum(1 for r in registros if r["competencia"] == comp)
        conn.execute(
            """
            INSERT INTO importacoes (empresa_id, competencia, arquivo, formato, registros)
            VALUES (?, ?, ?, 'IRPJ_CSLL', ?)
            """,
            (empresa_id, comp, str(arquivo), qtd),
        )

    conn.commit()
    return len(registros)


# ─── FUNÇÃO DE IMPORTAÇÃO ────────────────────────────────────────────────────

def importar_irpj_csll(caminho: Path, empresa_chave: str, conn: sqlite3.Connection) -> dict:
    """
    Importa a planilha ANUAL de IRPJ/CSLL para a empresa indicada (chave curta,
    ex. "mkb"/"gnileb" -- resolvida via `config.EMPRESAS`).

    Retorna {registros, competencias, empresa} ou {"erro": <msg>}.
    """
    emp = EMPRESAS.get(empresa_chave)
    if not emp:
        return {"erro": f"Empresa \"{empresa_chave}\" não encontrada em config.EMPRESAS."}

    resultado = parse_irpj_csll(caminho)
    if not resultado["registros"]:
        return {"erro": "Nenhuma linha de apuração válida encontrada na aba ANUAL."}

    qtd = salvar_irpj_csll(conn, resultado["registros"], emp["id"], caminho)

    return {
        "registros": qtd,
        "competencias": resultado["competencias"],
        "empresa": emp["sigla"],
    }


# ─── CLI DE TESTE ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from ingestion import get_conn, criar_schema, seed_empresas

    if len(sys.argv) < 3:
        print("Uso: python irpj_csll_parser.py <caminho_do_xlsx> <empresa (ex. mkb)>")
        sys.exit(1)

    caminho = Path(sys.argv[1])
    empresa_chave = sys.argv[2]
    conn = get_conn()
    criar_schema(conn)
    seed_empresas(conn)
    resultado = importar_irpj_csll(caminho, empresa_chave, conn)
    conn.close()
    print("\nResultado:", resultado)
