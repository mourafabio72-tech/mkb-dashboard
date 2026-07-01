"""
razao_parser.py -- MKB-Dashboard Sprint 4A
Parser do Razão Contábil Protheus (CT1 — relatório 12-00 Emissão do Razão Conta).

Estrutura do arquivo Razão:
  Linha 1 (header global): [cod_relatorio] [LOTE/SUB/DOC/LINHA] [HISTORICO] ...
  Linha "CONTA - X.X.X.XX.XX.XXX - DESCRICAO" → início de nova conta
  Linha "[CONTA] [DESCRICAO]" → separador de grupo (ignorar)
  Linha "[DD/MM/AAAA] [documento] [histórico] [c/partida] [filial] [cc] [_] [_] [débito] [crédito]" → lançamento
  Linha com "CONTA SEM MOVIMENTO NO PERIODO" → ignorar
"""

import re
import sqlite3
import datetime as _dt_mod
from pathlib import Path

import openpyxl

from config import EMPRESAS
from ingestion import parse_valor  # "1.234,56 C"/"1.234,56 D" -> float com sinal

# ─── PADRÕES ────────────────────────────────────────────────────────────────

# "CONTA - 4.1.1.03.07.007 - DESCRICAO" → captura código e descrição
_RE_CONTA = re.compile(r'CONTA\s*[-–]\s*([\d.]+)\s*[-–]\s*(.*)', re.IGNORECASE)

ABA_RAZAO = "12-00 - Emissao do Razao Conta"


# ─── PARSER NUMÉRICO ────────────────────────────────────────────────────────

def _parse_num(val) -> float:
    """Converte valor numérico ou texto para float."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(" ", "")
    if not s or s in ("-", "0,00", "0.00"):
        return 0.0
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _is_data(val) -> bool:
    """Retorna True se o valor for uma data (openpyxl converte cells de data para datetime)."""
    return isinstance(val, (_dt_mod.datetime, _dt_mod.date))


def _parse_saldo(val) -> float | None:
    """
    Converte a coluna "SALDO ATUAL" ('1.234,56 C'/'1.234,56 D') usando
    `parse_valor` (ingestion.py). Quando o saldo zera, o Protheus mostra
    "0,00" SEM sufixo C/D -- `parse_valor` exige o sufixo e retornaria None
    para esse caso, então tratamos esse formato explicitamente como zero.
    """
    if val is None:
        return None
    v = parse_valor(str(val))
    if v is not None:
        return v
    t = str(val).strip()
    if t in ("0,00", "0.00", "0"):
        return 0.0
    return None


def _extrair_data(val) -> tuple:
    """Extrai (data_str 'YYYY-MM-DD', competencia 'YYYY-MM') de um objeto datetime."""
    if isinstance(val, _dt_mod.datetime):
        return val.strftime("%Y-%m-%d"), val.strftime("%Y-%m")
    if isinstance(val, _dt_mod.date):
        return val.strftime("%Y-%m-%d"), val.strftime("%Y-%m")
    return None, None


# ─── PARSER PRINCIPAL ────────────────────────────────────────────────────────

def parse_razao(caminho: Path, empresa_id: int) -> tuple[list, dict]:
    """
    Lê o arquivo Razão Contábil do Protheus e retorna lista de dicts
    com os lançamentos individuais.

    Retorno: [{
        empresa_id, competencia, data_lanc, conta_cod,
        documento, historico, conta_partida, filial, centro_custo,
        debito, credito, valor
    }]
    """
    print(f"  Abrindo Razão: {caminho.name}")

    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)

    # Tenta encontrar a aba correta
    aba = None
    for nome in wb.sheetnames:
        if "razao" in nome.lower() or "razão" in nome.lower() or "12-00" in nome:
            aba = nome
            break
    if not aba:
        aba = wb.sheetnames[0]   # fallback: primeira aba

    ws = wb[aba]
    rows = list(ws.iter_rows(values_only=True))

    registros = []
    conta_atual = None
    desc_contas: dict[str, str] = {}
    n_sem_mov = 0

    for row in rows:
        col_a = str(row[0] if row[0] is not None else "").strip()
        col_b = str(row[1] if row[1] is not None else "").strip()
        col_c = str(row[2] if row[2] is not None else "").strip()
        col_d = str(row[3] if row[3] is not None else "").strip()  # conta partida
        col_e = str(row[4] if row[4] is not None else "").strip()  # filial
        col_f = str(row[5] if row[5] is not None else "").strip()  # centro de custo
        deb_raw  = row[8] if len(row) > 8 else None   # coluna I
        cred_raw = row[9] if len(row) > 9 else None   # coluna J
        saldo_raw = row[10] if len(row) > 10 else None  # coluna K -- "SALDO ATUAL"

        # ─── 1. Cabeçalho de conta individual ─────────────────────────────
        m_conta = _RE_CONTA.match(col_a)
        if m_conta:
            conta_atual = m_conta.group(1).strip()
            desc = m_conta.group(2).strip() if m_conta.group(2) else None
            if desc:
                desc_contas[conta_atual] = desc
            continue

        # ─── 2. Separadores de grupo e cabeçalhos repetidos (ignorar) ──────
        if col_a in ("CONTA", "TOTAL", "DATA") or col_b in ("DESCRICAO", "LOTE/SUB/DOC/LINHA"):
            continue

        # ─── 3. Linha de lançamento — col A é datetime nativo ─────────────
        if _is_data(row[0]) and conta_atual:

            # Ignora contas sem movimento
            if "SEM MOVIMENTO" in col_c.upper():
                n_sem_mov += 1
                continue

            data_str, competencia = _extrair_data(row[0])
            if not data_str:
                continue

            deb  = _parse_num(deb_raw)
            cred = _parse_num(cred_raw)

            # sinal DRE: receitas = crédito (+), despesas = débito (-)
            valor = cred - deb

            if deb == 0 and cred == 0:
                continue   # lançamento zerado — ignorar

            # Saldo acumulado da conta após este lançamento (coluna "SALDO
            # ATUAL", formato "1.234,56 C"/"1.234,56 D") -- usado por
            # Endividamento Tributário para saber o saldo devedor em
            # qualquer competência sem precisar reconstruir a partir do
            # saldo anterior (que este parser não captura). Não afeta DRE.
            saldo_atual = _parse_saldo(saldo_raw)

            registros.append({
                "empresa_id":   empresa_id,
                "competencia":  competencia,
                "data_lanc":    data_str,
                "conta_cod":    conta_atual,
                "documento":    col_b if col_b else None,
                "historico":    col_c[:200] if col_c else None,   # trunca
                "conta_partida": col_d if col_d else None,
                "filial":       col_e if col_e not in ("", "00") else None,
                "centro_custo": col_f if col_f else None,
                "debito":       deb,
                "credito":      cred,
                "valor":        valor,
                "saldo_atual":  saldo_atual,
            })

    wb.close()
    print(f"  {len(registros)} lançamentos | {n_sem_mov} contas sem movimento ignoradas")
    return registros, desc_contas


# ─── PERSISTÊNCIA ────────────────────────────────────────────────────────────

def salvar_razao(conn: sqlite3.Connection, registros: list, arquivo: Path, **kwargs) -> int:
    """
    Upsert dos lançamentos do Razão.
    Chave única: (empresa_id, data_lanc, documento, conta_cod).
    """
    if not registros:
        return 0

    # Garante que todo registro tem `parceiro_cod`/`saldo_atual` (campos
    # opcionais -- só o parser do Razão CT1 preenche `saldo_atual`; só os
    # importadores de CT2-detalhe preenchem `parceiro_cod`)
    for r in registros:
        r.setdefault("parceiro_cod", None)
        r.setdefault("saldo_atual", None)

    conn.executemany(
        """
        INSERT INTO razao
            (empresa_id, competencia, data_lanc, conta_cod, documento,
             historico, conta_partida, filial, centro_custo, debito, credito, valor, parceiro_cod, saldo_atual)
        VALUES
            (:empresa_id, :competencia, :data_lanc, :conta_cod, :documento,
             :historico, :conta_partida, :filial, :centro_custo, :debito, :credito, :valor, :parceiro_cod, :saldo_atual)
        ON CONFLICT (empresa_id, data_lanc, documento, conta_cod)
        DO UPDATE SET
            historico     = excluded.historico,
            conta_partida = excluded.conta_partida,
            filial        = excluded.filial,
            centro_custo  = excluded.centro_custo,
            debito        = excluded.debito,
            credito       = excluded.credito,
            valor         = excluded.valor,
            parceiro_cod  = excluded.parceiro_cod,
            saldo_atual   = excluded.saldo_atual
        """,
        registros,
    )

    # Atualiza tabela de contas (com descrição extraída do cabeçalho do Razão)
    desc_contas = kwargs.get("desc_contas", {})
    contas_params = []
    for r in registros:
        cod = r["conta_cod"]
        emp = r["empresa_id"]
        desc = desc_contas.get(cod)
        contas_params.append((cod, emp, desc))
    contas_uniq = {(c[0], c[1]): c[2] for c in contas_params}
    conn.executemany(
        """INSERT INTO contas (cod, empresa_id, descricao) VALUES (?, ?, ?)
           ON CONFLICT (cod, empresa_id) DO UPDATE SET
             descricao = COALESCE(NULLIF(excluded.descricao, ''), contas.descricao)""",
        [(cod, emp, desc) for (cod, emp), desc in contas_uniq.items()],
    )

    # Log de importação por competência
    comps = {}
    for r in registros:
        k = (r["empresa_id"], r["competencia"])
        comps[k] = comps.get(k, 0) + 1
    for (emp_id, comp), qtd in comps.items():
        conn.execute(
            "INSERT INTO importacoes (empresa_id, competencia, arquivo, registros) VALUES (?,?,?,?)",
            (emp_id, comp, str(arquivo), qtd)
        )

    conn.commit()
    return len(registros)


# ─── FUNÇÃO DE IMPORTAÇÃO ────────────────────────────────────────────────────

def importar_razao(caminho: Path, empresa_chave: str, conn: sqlite3.Connection) -> dict:
    """
    Importa um arquivo Razão Contábil para o banco.
    Retorna resumo: {registros, competencias}
    """
    emp = EMPRESAS.get(empresa_chave)
    if not emp:
        return {"erro": f"Empresa '{empresa_chave}' não encontrada"}

    registros, desc_contas = parse_razao(caminho, emp["id"])
    qtd = salvar_razao(conn, registros, caminho, desc_contas=desc_contas)
    competencias = sorted({r["competencia"] for r in registros})

    return {
        "registros":    qtd,
        "competencias": competencias,
        "empresa":      emp["sigla"],
    }


# ─── CLI DE TESTE ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from ingestion import get_conn, criar_schema, seed_empresas

    arq = Path(r"C:\Users\FabioMoura\BPS4 OUTSOURCING\Intranet BPS4 - Op. CONTABILIDADE"
               r"\04 - Grupo Markbuilding\00 - MKB\Razão\2023\MKB Razão 02 2023.xlsx")

    conn = get_conn()
    criar_schema(conn)
    seed_empresas(conn)

    res = importar_razao(arq, "mkb", conn)
    print(f"\nImportado: {res['registros']} lançamentos")
    print(f"Competências: {res['competencias']}")

    # Validação rápida
    row = conn.execute("""
        SELECT competencia, COUNT(*) as n, SUM(valor) as total
        FROM razao WHERE empresa_id=1
        GROUP BY competencia
    """).fetchall()
    print("\nResumo por competência (Razão):")
    for r in row:
        print(f"  {r[0]}  {r[1]} lançamentos  total={r[2]:,.2f}")

    conn.close()
