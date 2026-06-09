"""debug_razao.py -- inspeciona estrutura real do arquivo com openpyxl."""
import openpyxl
from pathlib import Path

arq = Path(r"C:\Users\FabioMoura\BPS4 OUTSOURCING\Intranet BPS4 - Op. CONTABILIDADE"
           r"\04 - Grupo Markbuilding\00 - MKB\Razão\2023\MKB Razão 02 2023.xlsx")

wb = openpyxl.load_workbook(arq, data_only=True, read_only=True)
print("Abas:", wb.sheetnames)

ws = wb.active
print(f"\nAba ativa: {ws.title}")
print(f"\nPrimeiras 40 linhas (colunas A-J):\n")

for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i > 40:
        break
    # Mostra tipo e valor de cada coluna
    cols = []
    for c, val in enumerate(row[:10], 1):
        if val is not None:
            cols.append(f"[{c}:{type(val).__name__}:{repr(val)[:30]}]")
        else:
            cols.append(f"[{c}:None]")
    print(f"L{i:02d}: {' '.join(cols)}")

wb.close()
